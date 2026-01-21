from fastapi import APIRouter, Depends, HTTPException, status
from typing import Optional
from pydantic import BaseModel
from datetime import datetime, date
from sqlmodel import select, func
from fastapi.responses import StreamingResponse
import io

from app.deps import DBSession, get_current_user
from app.models.models import Sale, SaleDetail, Product, Customer, CashRegisterSession, CashTransaction, Inventory, InventoryMovement, User
from app.auth.auth import RoleChecker

router = APIRouter()

# ========== SCHEMAS ==========
class FiltroReporte(BaseModel):
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None
    categoria_id: Optional[str] = None
    producto_id: Optional[str] = None
    cliente_id: Optional[str] = None
    usuario_id: Optional[str] = None

class ExportRequest(BaseModel):
    tipo: str  # ventas, inventario, caja, clientes
    filtro: FiltroReporte

# ========== ENDPOINTS DE REPORTES ==========
@router.post("/reportes/ventas", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Contador"]))])
async def reporte_ventas(
    filtro: FiltroReporte,
    db: DBSession,
    current_user: User = Depends(get_current_user)
):
    """
    Genera reporte de ventas con filtros
    """
    query = select(Sale)
    
    # Aplicar filtros
    if filtro.fecha_inicio:
        try:
            fecha_inicio = datetime.fromisoformat(filtro.fecha_inicio.replace('Z', '+00:00'))
            query = query.where(Sale.sale_date >= fecha_inicio)
        except:
            pass
    
    if filtro.fecha_fin:
        try:
            fecha_fin = datetime.fromisoformat(filtro.fecha_fin.replace('Z', '+00:00'))
            query = query.where(Sale.sale_date <= fecha_fin)
        except:
            pass
    
    if filtro.cliente_id:
        from uuid import UUID
        query = query.where(Sale.customer_id == UUID(filtro.cliente_id))
    
    sales = db.exec(query).all()
    
    # Calcular totales
    total_ventas = len(sales)
    total_ingresos = sum(sale.total_amount for sale in sales)
    total_descuentos = sum(sale.discount_amount for sale in sales)
    total_impuestos = sum(sale.tax_amount for sale in sales)
    
    # Ventas por día
    ventas_por_dia = {}
    for sale in sales:
        dia = sale.sale_date.date().isoformat()
        if dia not in ventas_por_dia:
            ventas_por_dia[dia] = {"cantidad": 0, "total": 0}
        ventas_por_dia[dia]["cantidad"] += 1
        ventas_por_dia[dia]["total"] += sale.total_amount
    
    return {
        "total_ventas": total_ventas,
        "total_ingresos": total_ingresos,
        "total_descuentos": total_descuentos,
        "total_impuestos": total_impuestos,
        "promedio_venta": total_ingresos / total_ventas if total_ventas > 0 else 0,
        "ventas_por_dia": ventas_por_dia,
        "periodo": {
            "inicio": filtro.fecha_inicio,
            "fin": filtro.fecha_fin
        }
    }

@router.post("/reportes/inventario", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Almacen"]))])
async def reporte_inventario(
    filtro: FiltroReporte,
    db: DBSession
):
    """
    Genera reporte de inventario actual
    """
    query = select(Inventory, Product).join(Product, Inventory.product_id == Product.id)
    
    if filtro.producto_id:
        from uuid import UUID
        query = query.where(Inventory.product_id == UUID(filtro.producto_id))
    
    results = db.exec(query).all()
    
    # Calcular estadísticas
    total_items = len(results)
    total_stock = sum(inv.quantity for inv, _ in results)
    productos_bajo_stock = sum(1 for inv, prod in results if inv.quantity <= prod.stock_min)
    valor_total = sum(inv.quantity * prod.cost_price for inv, prod in results)
    
    items = []
    for inv, prod in results:
        items.append({
            "producto_id": str(prod.id),
            "producto_nombre": prod.name,
            "sku": prod.sku,
            "cantidad": inv.quantity,
            "stock_min": prod.stock_min,
            "stock_max": prod.stock_max,
            "precio_costo": prod.cost_price,
            "valor_total": inv.quantity * prod.cost_price,
            "estado": "bajo" if inv.quantity <= prod.stock_min else "normal"
        })
    
    return {
        "total_items": total_items,
        "total_stock": total_stock,
        "productos_bajo_stock": productos_bajo_stock,
        "valor_total_inventario": valor_total,
        "items": items
    }

@router.post("/reportes/caja", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Contador"]))])
async def reporte_caja(
    filtro: FiltroReporte,
    db: DBSession
):
    """
    Genera reporte de movimientos de caja
    """
    query = select(CashRegisterSession)
    
    if filtro.fecha_inicio:
        try:
            fecha_inicio = datetime.fromisoformat(filtro.fecha_inicio.replace('Z', '+00:00'))
            query = query.where(CashRegisterSession.opening_date >= fecha_inicio)
        except:
            pass
    
    if filtro.fecha_fin:
        try:
            fecha_fin = datetime.fromisoformat(filtro.fecha_fin.replace('Z', '+00:00'))
            query = query.where(CashRegisterSession.opening_date <= fecha_fin)
        except:
            pass
    
    sessions = db.exec(query).all()
    
    # Calcular totales
    total_sesiones = len(sessions)
    total_ingresos = sum(s.opening_amount for s in sessions)
    total_cierres = sum(s.actual_closing_amount for s in sessions if s.actual_closing_amount)
    total_diferencias = sum(abs(s.difference) for s in sessions if s.difference)
    
    sesiones_data = []
    for session in sessions:
        sesiones_data.append({
            "id": str(session.id),
            "fecha_apertura": session.opening_date.isoformat(),
            "fecha_cierre": session.closing_date.isoformat() if session.closing_date else None,
            "monto_apertura": session.opening_amount,
            "monto_cierre": session.actual_closing_amount,
            "diferencia": session.difference,
            "status": session.status.value if hasattr(session.status, 'value') else session.status
        })
    
    return {
        "total_sesiones": total_sesiones,
        "total_ingresos": total_ingresos,
        "total_cierres": total_cierres,
        "total_diferencias": total_diferencias,
        "sesiones": sesiones_data,
        "periodo": {
            "inicio": filtro.fecha_inicio,
            "fin": filtro.fecha_fin
        }
    }

@router.post("/reportes/clientes", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Contador"]))])
async def reporte_clientes(
    filtro: FiltroReporte,
    db: DBSession
):
    """
    Genera reporte de clientes con compras
    """
    # Obtener todos los clientes
    customers = db.exec(select(Customer).where(Customer.is_active == True)).all()
    
    clientes_data = []
    total_clientes = len(customers)
    total_puntos = 0
    
    for customer in customers:
        # Contar compras del cliente
        sales_count = db.exec(
            select(func.count(Sale.id)).where(Sale.customer_id == customer.id)
        ).first() or 0
        
        # Total gastado
        total_spent = db.exec(
            select(func.sum(Sale.total_amount)).where(Sale.customer_id == customer.id)
        ).first() or 0.0
        
        total_puntos += customer.loyalty_points
        
        clientes_data.append({
            "id": str(customer.id),
            "nombre": f"{customer.first_name} {customer.last_name}",
            "documento": customer.document_number,
            "email": customer.email,
            "telefono": customer.phone,
            "puntos_fidelidad": customer.loyalty_points,
            "total_compras": sales_count,
            "total_gastado": float(total_spent) if total_spent else 0.0
        })
    
    # Ordenar por total gastado
    clientes_data.sort(key=lambda x: x["total_gastado"], reverse=True)
    
    return {
        "total_clientes": total_clientes,
        "total_puntos_sistema": total_puntos,
        "clientes": clientes_data,
        "top_clientes": clientes_data[:10]  # Top 10
    }

@router.post("/reportes/exportar/excel", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Contador"]))])
async def exportar_excel(
    request: ExportRequest,
    db: DBSession
):
    """
    Exporta reporte a Excel
    NOTA: Requiere openpyxl instalado: pip install openpyxl
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Librería openpyxl no instalada. Ejecute: pip install openpyxl"
        )
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    # Encabezados básicos
    ws['A1'] = 'Reporte Generado'
    ws['B1'] = datetime.now().isoformat()
    
    # Datos simulados (en producción llamaríamos al endpoint correspondiente)
    ws['A3'] = 'Tipo de Reporte'
    ws['B3'] = request.tipo
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_{request.tipo}.xlsx"}
    )

@router.post("/reportes/exportar/csv", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Contador"]))])
async def exportar_csv(
    request: ExportRequest,
    db: DBSession
):
    """
    Exporta reporte a CSV
    """
    import csv
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Encabezados
    writer.writerow(['Reporte', 'Fecha'])
    writer.writerow([request.tipo, datetime.now().isoformat()])
    writer.writerow([])
    
    # Datos según tipo
    writer.writerow(['Columna 1', 'Columna 2', 'Columna 3'])
    writer.writerow(['Dato 1', 'Dato 2', 'Dato 3'])
    
    # Convertir a bytes
    output.seek(0)
    content = output.getvalue().encode('utf-8')
    
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=reporte_{request.tipo}.csv"}
    )

@router.post("/reportes/exportar/pdf", tags=["Reportes"], dependencies=[Depends(RoleChecker(allowed_roles=["Administrador", "Contador"]))])
async def exportar_pdf(
    request: ExportRequest,
    db: DBSession
):
    """
    Exporta reporte a PDF
    NOTA: Requiere reportlab instalado: pip install reportlab
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Librería reportlab no instalada. Ejecute: pip install reportlab"
        )
    
    # Crear PDF en memoria
    output = io.BytesIO()
    c = canvas.Canvas(output, pagesize=letter)
    
    # Título
    c.setFont("Helvetica-Bold", 16)
    c.drawString(100, 750, f"Reporte de {request.tipo}")
    
    # Fecha
    c.setFont("Helvetica", 12)
    c.drawString(100, 730, f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Contenido simulado
    c.drawString(100, 700, "Datos del reporte aquí...")
    
    c.save()
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_{request.tipo}.pdf"}
    )
